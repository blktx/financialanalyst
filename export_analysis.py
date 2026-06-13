"""
Extended Analysis — Excel Export
Exports all 4 analysis areas to a formatted Excel report
with one tab per analysis area.

Run AFTER add_analysis.py has completed.
"""

import sqlite3
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

DB_PATH     = "/Users/beckylin/Desktop/2026/Financial Enginner/financial.db"
OUTPUT_PATH = "/Users/beckylin/Desktop/2026/Financial Enginner/Financial_Analysis_Report.xlsx"

conn = sqlite3.connect(DB_PATH)
print("✓ Connected to financial.db")

# ── Pull all 4 analysis datasets ─────────────────────────────────

fundamentals = pd.read_sql("""
    SELECT ticker, company, group_type, sector,
           ROUND(price, 2)          AS Price,
           ROUND(revenue_b, 1)      AS Revenue_B,
           ROUND(profit_margin, 1)  AS Profit_Margin_Pct,
           ROUND(roe, 1)            AS ROE_Pct,
           ROUND(debt_to_equity,1)  AS Debt_to_Equity,
           ROUND(current_ratio,2)   AS Current_Ratio,
           ROUND(pe_ratio, 1)       AS PE_Ratio,
           ROUND(free_cash_flow_b,1) AS FCF_B
    FROM stocks
    ORDER BY profit_margin DESC
""", conn)

sector = pd.read_sql("""
    SELECT sector, group_type,
           COUNT(*)                       AS Num_Stocks,
           ROUND(AVG(price), 2)           AS Avg_Price,
           ROUND(AVG(pe_ratio), 1)        AS Avg_PE,
           ROUND(AVG(dividend_yield), 2)  AS Avg_Yield_Pct,
           ROUND(AVG(profit_margin), 1)   AS Avg_Margin_Pct,
           ROUND(SUM(market_cap_b), 0)    AS Total_MktCap_B
    FROM stocks
    GROUP BY sector, group_type
    ORDER BY Total_MktCap_B DESC
""", conn)

trends = pd.read_sql("""
    SELECT ticker, company, group_type,
           ROUND(price, 2)           AS Price,
           ROUND(ma_50, 2)           AS MA_50,
           ROUND(ma_200, 2)          AS MA_200,
           ROUND(price_vs_ma50, 1)   AS Pct_vs_MA50,
           ROUND(week_52_high, 2)    AS High_52wk,
           ROUND(week_52_low, 2)     AS Low_52wk,
           ROUND(price_vs_52high, 1) AS Pct_from_52wk_High,
           CASE
               WHEN price_vs_ma50 > 0 AND price > ma_200 THEN 'Strong Uptrend'
               WHEN price_vs_ma50 > 0                    THEN 'Short Uptrend'
               WHEN price_vs_ma50 < 0 AND price < ma_200 THEN 'Downtrend'
               ELSE 'Mixed'
           END AS Trend_Signal
    FROM stocks
    ORDER BY Pct_vs_MA50 DESC
""", conn)

dividend_safety = pd.read_sql("""
    SELECT ticker, company,
           ROUND(price, 2)             AS Price,
           ROUND(dividend_yield, 2)    AS Yield_Pct,
           ROUND(payout_ratio, 1)      AS Payout_Ratio_Pct,
           ROUND(free_cash_flow_b, 2)  AS FCF_B,
           dividend_streak             AS Streak_Years,
           ROUND(roe, 1)               AS ROE_Pct,
           CASE
               WHEN payout_ratio < 60
                AND free_cash_flow_b > 0
                AND dividend_streak >= 5  THEN 'Safe'
               WHEN payout_ratio < 80
                AND free_cash_flow_b > 0  THEN 'Moderate'
               WHEN payout_ratio = 0
                OR  dividend_yield = 0    THEN 'No Dividend'
               ELSE 'At Risk'
           END AS Safety_Rating
    FROM stocks
    WHERE group_type = 'Dividend'
    ORDER BY Payout_Ratio_Pct ASC
""", conn)

conn.close()

# ── Write to Excel ────────────────────────────────────────────────
print("\nWriting Excel report...")

with pd.ExcelWriter(OUTPUT_PATH, engine="openpyxl") as writer:
    fundamentals.to_excel(writer,    sheet_name="1 Fundamentals",    index=False)
    sector.to_excel(writer,          sheet_name="2 Sector Breakdown", index=False)
    trends.to_excel(writer,          sheet_name="3 Trend Analysis",   index=False)
    dividend_safety.to_excel(writer, sheet_name="4 Dividend Safety",  index=False)

# ── Format the workbook ───────────────────────────────────────────
print("Formatting...")
wb = load_workbook(OUTPUT_PATH)

def style_sheet(ws, header_color, tab_color):
    ws.sheet_properties.tabColor = tab_color

    # Header row
    for cell in ws[1]:
        cell.font      = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        cell.fill      = PatternFill("solid", start_color=header_color, end_color=header_color)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin = Side(style="thin", color="CCCCCC")
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    ws.row_dimensions[1].height = 28

    # Data rows — zebra stripe
    for ri, row in enumerate(ws.iter_rows(min_row=2), start=2):
        bg = "F7F6F2" if ri % 2 == 0 else "FFFFFF"
        for cell in row:
            cell.font      = Font(name="Arial", size=10)
            cell.fill      = PatternFill("solid", start_color=bg, end_color=bg)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            thin = Side(style="thin", color="E0E0E0")
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Color code Safety_Rating column on dividend sheet
    if "Safety_Rating" in [c.value for c in ws[1]]:
        safety_col = None
        for cell in ws[1]:
            if cell.value == "Safety_Rating":
                safety_col = cell.column
        if safety_col:
            for row in ws.iter_rows(min_row=2):
                cell = row[safety_col - 1]
                if cell.value == "Safe":
                    cell.fill = PatternFill("solid", start_color="EAF3DE", end_color="EAF3DE")
                    cell.font = Font(name="Arial", size=10, color="3B6D11", bold=True)
                elif cell.value == "Moderate":
                    cell.fill = PatternFill("solid", start_color="FAEEDA", end_color="FAEEDA")
                    cell.font = Font(name="Arial", size=10, color="633806", bold=True)
                elif cell.value == "At Risk":
                    cell.fill = PatternFill("solid", start_color="FCEBEB", end_color="FCEBEB")
                    cell.font = Font(name="Arial", size=10, color="A32D2D", bold=True)

    # Color code Trend_Signal column
    if "Trend_Signal" in [c.value for c in ws[1]]:
        trend_col = None
        for cell in ws[1]:
            if cell.value == "Trend_Signal":
                trend_col = cell.column
        if trend_col:
            for row in ws.iter_rows(min_row=2):
                cell = row[trend_col - 1]
                if cell.value == "Strong Uptrend":
                    cell.fill = PatternFill("solid", start_color="EAF3DE", end_color="EAF3DE")
                    cell.font = Font(name="Arial", size=10, color="3B6D11", bold=True)
                elif cell.value == "Short Uptrend":
                    cell.fill = PatternFill("solid", start_color="E6F1FB", end_color="E6F1FB")
                    cell.font = Font(name="Arial", size=10, color="0C447C", bold=True)
                elif cell.value == "Downtrend":
                    cell.fill = PatternFill("solid", start_color="FCEBEB", end_color="FCEBEB")
                    cell.font = Font(name="Arial", size=10, color="A32D2D", bold=True)

    # Auto-fit column widths
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 28)

# Apply styles to each sheet
style_sheet(wb["1 Fundamentals"],    "3C3489", "7F77DD")
style_sheet(wb["2 Sector Breakdown"],"085041", "1D9E75")
style_sheet(wb["3 Trend Analysis"],  "0C447C", "4A90D9")
style_sheet(wb["4 Dividend Safety"], "633806", "BA7517")

wb.save(OUTPUT_PATH)
print(f"\n✓ Report saved to Desktop: Financial_Analysis_Report.xlsx")
print("  4 tabs — Fundamentals, Sector, Trends, Dividend Safety")
print("  Open it and explore your data!")
